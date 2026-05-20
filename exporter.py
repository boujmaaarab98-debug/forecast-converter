"""
exporter.py
Génère trois onglets dans un fichier Excel de sortie :
  1. PowerBI_Long   — table longue exploitable dans Power BI
  2. Format_Large   — tableau croisé une ligne par Ref
  3. Rapport_Controle — alertes et statistiques
"""

from __future__ import annotations

import pandas as pd
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from converter.warnings import ConversionWarnings


# ── Helpers de style ─────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
_ALT_FILL = PatternFill("solid", start_color="EBF3FB")
_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_ALERT_FILL = PatternFill("solid", start_color="FFF2CC")
_SECTION_FONT = Font(bold=True, name="Arial", size=10, color="1F4E79")


def _style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _style_data_row(ws, row_num: int, n_cols: int, alt: bool):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        if alt:
            cell.fill = _ALT_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = _BORDER
        cell.font = Font(name="Arial", size=9)


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


# ── Feuille 1 : Power BI Long ─────────────────────────────────────────────────

POWERBI_COLS = [
    "Client", "Source_File", "Sheet_Name", "File_Type",
    "Ref", "Désignation", "Division", "Cdc", "Reliquat",
    "Date_Mois", "Année", "Mois", "Quantité", "Type_Source", "Import_Date",
]


def _write_powerbi_sheet(ws, df: pd.DataFrame):
    ws.title = "PowerBI_Long"
    ws.freeze_panes = "A2"

    # Assurer les colonnes dans le bon ordre
    for col in POWERBI_COLS:
        if col not in df.columns:
            df[col] = ""
    df = df[POWERBI_COLS].copy()

    # Header
    for ci, col in enumerate(POWERBI_COLS, start=1):
        ws.cell(row=1, column=ci, value=col)
    _style_header_row(ws, 1, len(POWERBI_COLS))

    # Data
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
        _style_data_row(ws, ri, len(POWERBI_COLS), ri % 2 == 0)

    _auto_col_width(ws)


# ── Feuille 2 : Format Large ──────────────────────────────────────────────────

FIXED_COLS_LARGE = [
    "Client", "Ref", "Désignation", "Division", "Cdc",
    "cadence PIC A220", "besoin A220 2025", "codearticle",
    "comparaison code article", "Programme", "Coef avion", "cadence", "%", "GAMME",
]


def _build_wide_df(long_df: pd.DataFrame, client: str) -> pd.DataFrame:
    """Pivote la table longue en tableau large (une ligne par Ref, colonnes = mois)."""
    if long_df.empty:
        return pd.DataFrame()

    # Pivot
    pivot = long_df.pivot_table(
        index=["Client", "Ref", "Désignation", "Division", "Cdc"],
        columns="Date_Mois",
        values="Quantité",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    pivot.columns.name = None

    # Ajouter colonnes internes vides
    for col in ["cadence PIC A220", "besoin A220 2025", "codearticle",
                "comparaison code article", "Programme", "Coef avion", "cadence", "%", "GAMME"]:
        pivot[col] = ""

    # Réordonner : colonnes fixes d'abord, puis mois triés
    date_cols = sorted([c for c in pivot.columns if _is_date_col(c)])
    ordered = FIXED_COLS_LARGE + date_cols

    for col in ordered:
        if col not in pivot.columns:
            pivot[col] = ""

    return pivot[ordered]


def _is_date_col(c: str) -> bool:
    import re
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", str(c)))


def _write_wide_sheet(ws, wide_df: pd.DataFrame):
    ws.title = "Format_Large"
    ws.freeze_panes = "C2"

    if wide_df.empty:
        ws.cell(row=1, column=1, value="Aucune donnée à afficher.")
        return

    cols = list(wide_df.columns)

    for ci, col in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value=col)
    _style_header_row(ws, 1, len(cols))

    # Distinguer colonnes fixes vs date pour coloration
    n_fixed = len(FIXED_COLS_LARGE)

    for ri, row in enumerate(wide_df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val if val != 0 else "")
            cell.border = _BORDER
            cell.font = Font(name="Arial", size=9)
            if ri % 2 == 0:
                cell.fill = _ALT_FILL

    _auto_col_width(ws)


# ── Feuille 3 : Rapport de Contrôle ──────────────────────────────────────────

def _write_rapport_sheet(ws, warn: ConversionWarnings):
    ws.title = "Rapport_Controle"

    rows = warn.to_dict_list()

    ws.cell(row=1, column=1, value="Catégorie")
    ws.cell(row=1, column=2, value="Détail")
    _style_header_row(ws, 1, 2)

    for ri, item in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=item["Catégorie"]).font = _SECTION_FONT
        ws.cell(row=ri, column=2, value=item["Détail"])
        ws.cell(row=ri, column=2).fill = _ALERT_FILL
        ws.cell(row=ri, column=1).border = _BORDER
        ws.cell(row=ri, column=2).border = _BORDER

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 55


# ── Export principal ──────────────────────────────────────────────────────────

def export_all(long_df: pd.DataFrame, warn: ConversionWarnings,
               output_path: str | Path, client: str):
    """
    Génère le fichier Excel final avec 3 onglets.
    Retourne le chemin du fichier créé.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Supprimer la feuille par défaut
    default = wb.active
    wb.remove(default)

    ws_pb = wb.create_sheet("PowerBI_Long")
    _write_powerbi_sheet(ws_pb, long_df.copy() if not long_df.empty else pd.DataFrame(columns=POWERBI_COLS))

    ws_wide = wb.create_sheet("Format_Large")
    wide_df = _build_wide_df(long_df, client) if not long_df.empty else pd.DataFrame()
    _write_wide_sheet(ws_wide, wide_df)

    ws_rpt = wb.create_sheet("Rapport_Controle")
    _write_rapport_sheet(ws_rpt, warn)

    wb.save(output_path)
    return output_path
